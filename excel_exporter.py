#!/usr/bin/env python3
"""
Excel Exporter
Handles exporting results to Excel with formatting
"""

import os
import subprocess
import sys
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Fix Windows encoding issues
if sys.platform == "win32":
    import codecs
    import io
    try:
        # Check if stdout has a buffer attribute (not in subprocess environments)
        if hasattr(sys.stdout, 'buffer'):
            sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer)
            sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer)
    except (AttributeError, OSError):
        # Fallback for subprocess environments or different Windows setups
        pass  # Keep original stdout/stderr


class ExcelExporter:
    def __init__(self):
        # Automatically detect correct Python command based on OS
        self.python_cmd = self._get_python_command()
        print(f"🐍 Using Python command: {self.python_cmd}")
    
    def _get_python_command(self):
        """Get the correct Python command for current OS"""
        if sys.platform == "win32":
            # Test different Python commands on Windows
            test_commands = ["python", "python3", "py"]
            for cmd in test_commands:
                try:
                    result = subprocess.run([cmd, "--version"], capture_output=True, text=True)
                    if result.returncode == 0:
                        return cmd
                except:
                    continue
            return "python"  # Default fallback
        else:
            return "python3"  # Unix/Linux/macOS
    
    def export_config_report_to_xlsx(self, config_data: Dict) -> str:
        """
        Export report based on saved configuration to XLSX file.
        Uses config parameters to run daily_report and deals_categorizer, 
        then exports their output to Excel.
        
        Args:
            config_data: Dictionary containing configuration parameters
            
        Returns:
            str: Filename of the exported Excel file
        """
        try:
            # Get export filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            config_name = config_data.get('name', 'config_report')
            database = config_data.get('database', 'unknown')
            
            # Create filename
            filename = f"{database}_{config_name}_{timestamp}.xlsx"
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            self._create_config_summary_sheet(summary_sheet, config_data)
            
            # Prepare commands based on configuration
            daily_report_data = None
            deals_categorizer_data = None
            
            # Build daily_report command and run it
            if config_data.get('database'):
                daily_report_cmd = self._build_daily_report_command(config_data)
                print(f"[R] Running daily report command: {' '.join(daily_report_cmd[:7])}...")
                daily_report_data = self._run_command_and_capture_output(daily_report_cmd)
                
                if daily_report_data:
                    print(f"✅ Daily report data received ({len(daily_report_data)} characters)")
                else:
                    print("⚠️ No daily report data received")
            
            # Build deals_categorizer command and run it  
            if config_data.get('database'):
                deals_categorizer_cmd = self._build_deals_categorizer_command(config_data)
                print(f"[C] Running deals categorizer command: {' '.join(deals_categorizer_cmd[:7])}...")
                deals_categorizer_data = self._run_command_and_capture_output(deals_categorizer_cmd)
                
                if deals_categorizer_data:
                    print(f"✅ Deals categorizer data received ({len(deals_categorizer_data)} characters)")
                else:
                    print("⚠️ No deals categorizer data received")
            
            # Create Daily Report sheet (always create, even if no data)
            daily_sheet = wb.create_sheet(title="Daily_Report")
            self._create_config_report_sheet(daily_sheet, daily_report_data, "Daily Report")
            
            # Create Deals Categorizer sheet (always create, even if no data)
            deals_sheet = wb.create_sheet(title="Deals_Categorizer")
            self._create_config_deals_sheet(deals_sheet, deals_categorizer_data, "Deals Categorizer", config_data)
            
            # Save the workbook
            wb.save(filename)
            
            print(f"✓ Config report exported to: {filename}")
            print(f"📁 File saved in: {os.path.abspath(filename)}")
            
            # Show sheet summary
            sheet_count = len(wb.sheetnames)
            print(f"[R] Contains {sheet_count} sheets:")
            for sheet_name in wb.sheetnames:
                print(f"   - {sheet_name}")
            
            return filename
            
        except Exception as e:
            print(f"❌ Error exporting config report: {e}")
            return None
    
    def _create_config_summary_sheet(self, ws, config_data: Dict):
        """Create summary sheet for configuration-based report"""
        # Title
        ws.append(["Configuration Report Summary"])
        ws.append([])
        
        # Basic info
        ws.append(["Configuration Name", config_data.get('name', 'N/A')])
        ws.append(["Database", config_data.get('database', 'N/A')])
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        # Groups information
        groups = config_data.get('groups', [])
        if groups:
            ws.append(["Groups", f"{len(groups)} selected"])
            ws.append([])
            ws.append(["Selected Groups:"])
            for group in groups:
                ws.append(["", group])
        else:
            ws.append(["Groups", "All"])
        
        # Login range
        ws.append([])
        min_login = config_data.get('min_login')
        max_login = config_data.get('max_login')
        min_login_str = f"{min_login:,}" if min_login is not None else "N/A"
        max_login_str = f"{max_login:,}" if max_login is not None else "N/A"
        ws.append(["Login Range", f"{min_login_str} - {max_login_str}"])
        
        # Time period
        if config_data.get('start_date') and config_data.get('end_date'):
            ws.append(["Date Range", f"{config_data['start_date']} to {config_data['end_date']}"])
        
        # Other parameters
        if config_data.get('min_profit') is not None:
            ws.append(["Min Profit", f"{config_data['min_profit']:,.2f}"])
        if config_data.get('max_profit') is not None:
            ws.append(["Max Profit", f"{config_data['max_profit']:,.2f}"])
        
        # Agent and ZIP information
        if config_data.get('agent'):
            ws.append(["Agent", config_data['agent']])
        if config_data.get('zip'):
            ws.append(["ZIP", config_data['zip']])
        
        # Style the summary
        self._style_simple_summary_sheet(ws)
    
    def _build_daily_report_command(self, config_data: Dict) -> List[str]:
        """Build daily_report command from configuration"""
        cmd = [self.python_cmd, "daily_report.py"]
        
        # Add database
        if config_data.get('database'):
            cmd.extend(["--database", config_data['database']])
        
        # Add date if specified
        if config_data.get('date'):
            cmd.extend(["--date", config_data['date']])
        elif config_data.get('start_date'):
            cmd.extend(["--date", config_data['start_date']])
        
        # Add limit - ensure we use the same limit as the config
        limit = config_data.get('limit')
        if limit and limit > 0:
            cmd.extend(["--limit", str(limit)])
        else:
            # Use a reasonable default if no limit specified
            cmd.extend(["--limit", "100"])
        
        # Add groups filter
        if config_data.get('groups'):
            cmd.extend(["--groups"] + config_data['groups'])
        
        # Add login range filters
        if config_data.get('min_login'):
            cmd.extend(["--min-login", str(config_data['min_login'])])
        
        if config_data.get('max_login'):
            cmd.extend(["--max-login", str(config_data['max_login'])])
        
        # Add profit filters
        if config_data.get('min_profit') is not None:
            cmd.extend(["--min-profit", str(config_data['min_profit'])])
        
        if config_data.get('max_profit') is not None:
            cmd.extend(["--max-profit", str(config_data['max_profit'])])
        
        # Add agent filter
        if config_data.get('agent'):
            cmd.extend(["--agent", config_data['agent']])
        
        # Add zip filter
        if config_data.get('zip'):
            cmd.extend(["--zip", config_data['zip']])
        
        return cmd
    
    def _build_deals_categorizer_command(self, config_data: Dict) -> List[str]:
        """Build deals_categorizer command from configuration"""
        cmd = [self.python_cmd, "deals_categorizer.py"]
        
        # Add database
        if config_data.get('database'):
            cmd.extend(["--database", config_data['database']])
        
        # Add year - use config date or current year
        if config_data.get('start_date'):
            # Extract year from start_date
            try:
                year = int(config_data['start_date'][:4])
                cmd.extend(["--year", str(year)])
            except (ValueError, IndexError):
                cmd.extend(["--year", str(datetime.now().year)])
        else:
            cmd.extend(["--year", str(datetime.now().year)])
        
        # Add limit - use config limit or reasonable default
        limit = config_data.get('limit')
        if limit and limit > 0:
            cmd.extend(["--limit", str(limit)])
        else:
            # Use reasonable default for deals
            cmd.extend(["--limit", "100"])
        
        # Add monthly flag to get detailed data
        cmd.append("--monthly")
        
        # Note: deals_categorizer.py doesn't currently support:
        # - groups filter
        # - login range filter  
        # - profit range filter
        # - agent filter
        # - zip filter
        # These filters would need to be added to deals_categorizer.py
        
        return cmd
    
    def _run_command_and_capture_output(self, command: List[str]) -> str:
        """Run a command and capture its output"""
        try:
            print(f"[>>] Executing: {' '.join(command[:3])}...")
            
            # Try different encoding strategies based on platform
            if sys.platform == "win32":
                # Windows: Try CP1252 first, then UTF-8 with error handling
                encodings_to_try = ['cp1252', 'utf-8', 'latin-1']
            else:
                # Unix-like: Try UTF-8 first
                encodings_to_try = ['utf-8', 'latin-1']
            
            result = None
            for encoding in encodings_to_try:
                try:
                    result = subprocess.run(
                        command,
                        capture_output=True,
                        text=True,
                        encoding=encoding,
                        errors='replace',  # Replace problematic characters
                        timeout=120  # 2 minutes timeout
                    )
                    break  # Success with this encoding
                except (UnicodeDecodeError, UnicodeError):
                    continue  # Try next encoding
            
            if result is None:
                print(f"❌ Failed to decode command output with any encoding")
                return None
            
            if result.returncode == 0:
                print(f"✅ Command completed successfully")
                if result.stdout and result.stdout.strip():
                    print(f"📝 Output length: {len(result.stdout)} characters")
                    return result.stdout
                else:
                    print("⚠️ Command completed but no output received")
                    return None
            else:
                print(f"❌ Command failed with return code {result.returncode}")
                if result.stderr:
                    print(f"📝 Error output: {result.stderr[:500]}...")
                if result.stdout:
                    print(f"📝 Standard output: {result.stdout[:500]}...")
                return None
                
        except subprocess.TimeoutExpired:
            print(f"⏰ Command timed out after 2 minutes: {' '.join(command[:3])}")
            return None
        except Exception as e:
            print(f"❌ Error running command: {e}")
            return None
    
    def _create_config_report_sheet(self, ws, output_data: str, report_title: str):
        """Create report sheet from command output"""
        # Add title
        ws.append([f"{report_title}"])
        ws.append([])
        
        # Handle None output
        if output_data is None:
            ws.append(["No data available - command failed or produced no output"])
            return
        
        # Parse the output data
        parsed_data = self._parse_command_output(output_data)
        
        # Add all data to worksheet
        if parsed_data:
            # Sort by Net P/L column if it exists
            if len(parsed_data) > 1:  # Skip header row
                header_row = parsed_data[0]
                data_rows = parsed_data[1:]
                
                # Find Net P/L column index
                net_pl_index = None
                for i, header in enumerate(header_row):
                    if 'Net P/L' in str(header):
                        net_pl_index = i
                        break
                
                # Sort by Net P/L if found (ascending - from low to high)
                if net_pl_index is not None:
                    try:
                        def get_net_pl_value(row):
                            if len(row) <= net_pl_index or not row[net_pl_index]:
                                return 0
                            try:
                                # Clean the value and convert to float
                                cleaned = str(row[net_pl_index]).replace('$', '').replace(',', '').strip()
                                return float(cleaned)
                            except (ValueError, TypeError):
                                return 0
                        
                        data_rows.sort(key=get_net_pl_value, reverse=False)
                    except (ValueError, IndexError):
                        # If sorting fails, just keep original order
                        pass
                
                # Rebuild parsed_data with sorted rows
                parsed_data = [header_row] + data_rows
            
            for row_data in parsed_data:
                ws.append(row_data)
            
            # Style the sheet
            self._style_clean_data_sheet(ws, report_title)
        else:
            # If no structured data found, show a message
            ws.append(["No structured data found in the report output"])
            print("⚠️ No structured data found in the report output")
    
    def _create_config_deals_sheet(self, ws, output_data: str, report_title: str, config_data: Dict = None):
        """Create deals sheet from deals_categorizer output"""
        # Add title
        ws.append([f"{report_title}"])
        ws.append([])
        
        # Handle None output
        if output_data is None:
            ws.append(["No data available - command failed or produced no output"])
            return
        
        # Parse the deals data
        deals_data = self._parse_deals_categorizer_output(output_data)
        
        if deals_data:
            # Add headers - include agent and zip columns
            headers = ["Login", "Year", "Month", "Deal ID", "Category", "Profit", "Comment", "Date", "Agent", "ZIP"]
            ws.append(headers)
            
            # Add deal data
            for deal in deals_data:
                ws.append([
                    deal.get('login', ''),
                    deal.get('year', ''),
                    deal.get('month_name', ''),
                    deal.get('deal_id', ''),
                    deal.get('category', ''),
                    deal.get('profit', 0.0),
                    deal.get('comment', ''),
                    deal.get('date', ''),
                    deal.get('agent', ''),
                    deal.get('zip_code', '')
                ])
            
            # Style the deals sheet
            self._style_deals_data_sheet(ws, report_title)
        else:
            # If no deals found, show a message
            ws.append(["No deal data found"])
    
    def _filter_deals_by_config(self, deals_data: List[Dict], config_data: Dict) -> List[Dict]:
        """Filter deals data based on config parameters"""
        if not deals_data:
            return []
        
        # Get all unique logins from deals
        logins = set()
        for deal in deals_data:
            login = deal.get('login')
            if login:
                try:
                    logins.add(int(login))
                except (ValueError, TypeError):
                    continue
        
        # Get login-group mapping if groups filter is specified
        login_group_mapping = {}
        if config_data.get('groups') and logins:
            database = config_data.get('database', 'mt5gn_live')
            login_group_mapping = self._get_login_group_mapping(database, list(logins))
        
        filtered_deals = []
        
        for deal in deals_data:
            # Apply login range filter
            login = deal.get('login')
            if login:
                try:
                    login_num = int(login)
                    min_login = config_data.get('min_login')
                    max_login = config_data.get('max_login')
                    
                    if min_login and login_num < min_login:
                        continue
                    if max_login and login_num > max_login:
                        continue
                    
                    # Apply groups filter
                    if config_data.get('groups') and login_group_mapping:
                        user_group = login_group_mapping.get(login_num)
                        if user_group and user_group not in config_data['groups']:
                            continue
                        elif not user_group:
                            # Skip if we couldn't find the group for this login
                            continue
                    
                except (ValueError, TypeError):
                    # Skip if login is not a number
                    continue
            
            # Apply profit range filter
            profit = deal.get('profit')
            if profit is not None:
                try:
                    profit_num = float(profit)
                    min_profit = config_data.get('min_profit')
                    max_profit = config_data.get('max_profit')
                    
                    if min_profit is not None and profit_num < min_profit:
                        continue
                    if max_profit is not None and profit_num > max_profit:
                        continue
                except (ValueError, TypeError):
                    # Skip if profit is not a number
                    continue
            
            # If all filters pass, add the deal
            filtered_deals.append(deal)
        
        return filtered_deals
    
    def export_results_to_xlsx(self, results: List[Dict], config: Dict) -> str:
        """Export results to XLSX file with organized sheets and return filename"""
        try:
            # Get export filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            database = config.get('database', 'unknown')
            report_type = config.get('report_type', 'report')
            
            # Create more descriptive filename
            filename_parts = [database, report_type, timestamp]
            if config.get('groups'):
                filename_parts.insert(-1, f"{len(config['groups'])}groups")
            
            filename = "_".join(filename_parts) + ".xlsx"
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            self._create_summary_sheet(summary_sheet, results, config)
            
            # Organize results by type
            daily_report_results = []
            deals_categorizer_results = []
            
            for result in results:
                if 'daily_report.py' in result['command']:
                    daily_report_results.append(result)
                elif 'deals_categorizer.py' in result['command']:
                    deals_categorizer_results.append(result)
            
            # Create Daily Report sheet
            if daily_report_results:
                daily_sheet = wb.create_sheet(title="Daily_Report")
                self._create_report_sheet(daily_sheet, daily_report_results, "Daily Report")
            
            # Create Deals Categorizer sheet with deal-by-deal data
            if deals_categorizer_results:
                deals_sheet = wb.create_sheet(title="Deals_Categorizer")
                self._create_deals_detailed_sheet(deals_sheet, deals_categorizer_results, "Deals Categorizer")
            
            # If no specific type identified, create generic sheets
            other_results = [r for r in results if 'daily_report.py' not in r['command'] and 'deals_categorizer.py' not in r['command']]
            if other_results:
                other_sheet = wb.create_sheet(title="Other_Reports")
                self._create_report_sheet(other_sheet, other_results, "Other Reports")
            
            # Save the workbook
            wb.save(filename)
            
            print(f"✓ Results exported to: {filename}")
            print(f"📁 File saved in: {os.path.abspath(filename)}")
            
            # Show sheet summary
            sheet_count = len(wb.sheetnames)
            print(f"📊 Contains {sheet_count} sheets:")
            for sheet_name in wb.sheetnames:
                print(f"   - {sheet_name}")
            
            return filename
            
        except Exception as e:
            print(f"❌ Error exporting results: {e}")
            return None
    
    def _create_summary_sheet(self, ws, results: List[Dict], config: Dict):
        """Create a simple summary sheet with basic information"""
        # Title
        ws.append(["Report Summary"])
        ws.append([])
        
        # Basic info
        ws.append(["Database", config.get('database', 'N/A')])
        ws.append(["Report Type", config.get('report_type', 'N/A')])
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        if config.get('groups'):
            ws.append(["Groups", f"{len(config['groups'])} selected"])
        else:
            ws.append(["Groups", "All"])
        
        min_login = config.get('min_login')
        max_login = config.get('max_login')
        min_login_str = f"{min_login:,}" if min_login is not None else "N/A"
        max_login_str = f"{max_login:,}" if max_login is not None else "N/A"
        ws.append(["Login Range", f"{min_login_str} - {max_login_str}"])
        
        # Style the summary
        self._style_simple_summary_sheet(ws)
    
    def _create_report_sheet(self, ws, results: List[Dict], report_title: str):
        """Create a clean report sheet with only data and headers"""
        # Add title
        ws.append([f"{report_title}"])
        ws.append([])
        
        # Process each result and combine all data
        all_data = []
        
        for result in results:
            # Process the output data
            output_data = self._parse_command_output(result['output'])
            
            if output_data:
                all_data.extend(output_data)
        
        # Add all data to worksheet
        if all_data:
            # Sort by Net P/L column if it exists
            if len(all_data) > 1:  # Skip header row
                header_row = all_data[0]
                data_rows = all_data[1:]
                
                # Find Net P/L column index
                net_pl_index = None
                for i, header in enumerate(header_row):
                    if 'Net P/L' in str(header):
                        net_pl_index = i
                        break
                
                # Sort by Net P/L if found (ascending - from low to high)
                if net_pl_index is not None:
                    try:
                        def get_net_pl_value(row):
                            if len(row) <= net_pl_index or not row[net_pl_index]:
                                return 0
                            try:
                                # Clean the value and convert to float
                                cleaned = str(row[net_pl_index]).replace('$', '').replace(',', '').strip()
                                return float(cleaned)
                            except (ValueError, TypeError):
                                return 0
                        
                        data_rows.sort(key=get_net_pl_value, reverse=False)
                    except (ValueError, IndexError):
                        # If sorting fails, just keep original order
                        pass
                
                # Rebuild all_data with sorted rows
                all_data = [header_row] + data_rows
            
            for row_data in all_data:
                ws.append(row_data)
            
            # Style the sheet
            self._style_clean_data_sheet(ws, report_title)
        else:
            # If no structured data found, show a message
            ws.append(["No structured data found in the report output"])
    
    def _create_deals_detailed_sheet(self, ws, results: List[Dict], report_title: str):
        """Create a detailed deals sheet with deal-by-deal data from deals_categorizer"""
        # Add title
        ws.append([f"{report_title}"])
        ws.append([])
        
        # Process deals categorizer results
        all_deals = []
        
        for result in results:
            # Parse the deals data from the command output
            deals_data = self._parse_deals_categorizer_output(result['output'])
            all_deals.extend(deals_data)
        
        if all_deals:
            # Add headers - include agent and zip columns
            headers = ["Login", "Year", "Month", "Deal ID", "Category", "Profit", "Comment", "Date", "Agent", "ZIP"]
            ws.append(headers)
            
            # Add deal data
            for deal in all_deals:
                ws.append([
                    deal.get('login', ''),
                    deal.get('year', ''),
                    deal.get('month_name', ''),
                    deal.get('deal_id', ''),
                    deal.get('category', ''),
                    deal.get('profit', 0.0),
                    deal.get('comment', ''),
                    deal.get('date', ''),
                    deal.get('agent', ''),
                    deal.get('zip_code', '')
                ])
            
            # Style the deals sheet
            self._style_deals_data_sheet(ws, report_title)
        else:
            # If no deals found, show a message
            ws.append(["No deal-by-deal data found in the report output"])
    
    def _parse_command_output(self, output: str) -> List[List[str]]:
        """Parse command output to extract clean tabular data"""
        if not output:
            return []
        
        lines = output.strip().split('\n')
        parsed_data = []
        
        for line in lines:
            if not line:
                continue
            
            line = line.strip()
            if not line:
                continue
            
            # Skip unwanted lines from reports
            if (line.startswith('Groups filter:') or 
                line.startswith('Row ') or
                line.startswith('Login=') or
                line.startswith('Category=') or
                line.startswith('Profit=') or
                line.startswith('Count=') or
                'Row 1' in line or
                'Row 2' in line or
                'Row 3' in line or
                'Row 4' in line or
                'Row 5' in line or
                'Row 6' in line or
                'Row 7' in line or
                'Row 8' in line or
                'Row 9' in line or
                line.startswith('Row') or
                line.startswith('Groups filter')):
                continue
            
            # Look for pipe-separated table data (primary format)
            if '|' in line:
                cells = [cell.strip() if cell else '' for cell in line.split('|')]
                cells = [cell for cell in cells if cell]  # Remove empty cells
                if cells and len(cells) > 1:
                    # Skip separator lines (all dashes/spaces)
                    if not all(c.replace('-', '').replace(' ', '') == '' for c in cells):
                        # Apply minimal cell cleaning
                        cells = self._clean_cell_data_minimal(cells)
                        parsed_data.append(cells)
            
            # Look for CSV format data as fallback
            elif ',' in line and len(line.split(',')) > 2:
                cells = [cell.strip() if cell else '' for cell in line.split(',')]
                if len(cells) > 2:  # Only accept rows with more than 2 columns
                    cells = self._clean_cell_data_minimal(cells)
                    parsed_data.append(cells)
        
        return parsed_data
    
    def _parse_deals_categorizer_output(self, output: str) -> List[Dict]:
        """Parse deals categorizer output to extract deal-by-deal data"""
        if not output:
            return []
            
        lines = output.strip().split('\n')
        deals = []
        in_monthly_table = False
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Look for the monthly deals table header
            if "Monthly Deals by Login" in line:
                in_monthly_table = True
                continue
            
            # Skip table borders and separators
            if line.startswith('=') or line.startswith('-') or line.startswith('+'):
                continue
            
            # Parse table data if we're in the monthly deals section
            if in_monthly_table and '|' in line:
                cells = [cell.strip() for cell in line.split('|')]
                # Remove empty cells from edges
                cells = [cell for cell in cells if cell]
                
                # Skip header row and separator rows
                if (len(cells) >= 10 and 
                    cells[0] != 'Login' and 
                    not all(c.replace('-', '').replace(' ', '') == '' for c in cells)):
                    
                    try:
                        # Parse the deal data - now including agent and zip
                        deal = {
                            'login': int(cells[0]) if cells[0].isdigit() else cells[0],
                            'year': int(cells[1]) if cells[1].isdigit() else cells[1],
                            'month_name': cells[2],
                            'deal_id': int(cells[3]) if cells[3].isdigit() else cells[3],
                            'category': cells[4],
                            'profit': self._clean_numeric_value(cells[5]),
                            'comment': cells[6],
                            'date': cells[7] if len(cells) > 7 else '',
                            'agent': cells[8] if len(cells) > 8 else '',
                            'zip_code': cells[9] if len(cells) > 9 else ''
                        }
                        deals.append(deal)
                    except (ValueError, IndexError):
                        # Skip malformed rows
                        continue
        
        return deals
    
    def _is_numeric_like(self, text: str) -> bool:
        """Check if text looks like a numeric value"""
        try:
            # Remove common currency symbols and formatting
            cleaned = text.replace('$', '').replace(',', '').replace('%', '')
            float(cleaned)
            return True
        except ValueError:
            return False
    
    def _clean_numeric_value(self, value: str):
        """Clean and convert string values to proper numeric types"""
        if not isinstance(value, str):
            return value
        
        # Remove common currency symbols and formatting
        cleaned = value.replace('$', '').replace(',', '').replace('%', '').strip()
        
        # Try to convert to number
        try:
            # Check if it's an integer
            if '.' not in cleaned and cleaned.isdigit():
                return int(cleaned)
            # Check if it's a float
            elif cleaned.replace('.', '').replace('-', '').isdigit():
                return float(cleaned)
            else:
                return value  # Return original if not a number
        except:
            return value  # Return original if conversion fails
    
    def _style_simple_summary_sheet(self, ws):
        """Style the simple summary sheet"""
        # Title styling
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E4A75", end_color="2E4A75", fill_type="solid")
        
        # Data styling
        for row in range(3, ws.max_row + 1):
            # Bold the labels in column A
            label_cell = ws.cell(row=row, column=1)
            label_cell.font = Font(bold=True)
            
            # Style both columns
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
    
    def _style_clean_data_sheet(self, ws, report_title: str):
        """Style the clean data sheet with headers and proper number formatting"""
        if ws.max_row < 3:
            return
        
        # Title styling
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E4A75", end_color="2E4A75", fill_type="solid")
        
        # Find the maximum column count
        max_col = 1
        for row in range(3, ws.max_row + 1):
            row_cells = [cell for cell in ws[row] if cell.value is not None]
            if row_cells:
                max_col = max(max_col, len(row_cells))
        
        # Style the first data row as headers (row 3)
        if ws.max_row >= 3:
            for col in range(1, max_col + 1):
                header_cell = ws.cell(row=3, column=col)
                header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                header_cell.fill = PatternFill(start_color="5B7FA6", end_color="5B7FA6", fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style data rows with alternating colors and proper number formatting
        for row in range(4, ws.max_row + 1):
            fill_color = "F2F2F2" if row % 2 == 0 else "FFFFFF"
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Convert numeric values and format them
                if cell.value and isinstance(cell.value, str):
                    cleaned_value = self._clean_numeric_value(cell.value)
                    if cleaned_value is not None:
                        cell.value = cleaned_value
                        # Format based on the type of number and column
                        if isinstance(cleaned_value, float):
                            cell.number_format = '#,##0.00'
                        elif isinstance(cleaned_value, int):
                            # Don't add thousands separator for Login column (column 1)
                            if col == 1:
                                cell.number_format = '0'  # No thousands separator for Login
                            else:
                                cell.number_format = '#,##0'
        
        # Add borders to all data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(3, ws.max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        # Auto-adjust column widths with specific handling for Group column
        for col in range(1, max_col + 1):
            max_width = 0
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_width = max(max_width, len(str(cell.value)))
            
            # Set column width (with some padding)
            column_letter = ws.cell(row=1, column=col).column_letter
            
            # Special handling for Group column (typically column 3)
            if col == 3:
                # Ensure Group column has adequate width (minimum 20, maximum 60)
                ws.column_dimensions[column_letter].width = min(max(max_width + 5, 20), 60)
            else:
                ws.column_dimensions[column_letter].width = min(max_width + 3, 50)
    
    def _style_deals_data_sheet(self, ws, report_title: str):
        """Style the deals data sheet with proper formatting"""
        if ws.max_row < 3:
            return
        
        # Title styling
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E4A75", end_color="2E4A75", fill_type="solid")
        
        # Find the maximum column count
        max_col = 10  # We now have 10 columns for deals (added agent and zip)
        
        # Style the header row (row 3)
        if ws.max_row >= 3:
            for col in range(1, max_col + 1):
                header_cell = ws.cell(row=3, column=col)
                header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                header_cell.fill = PatternFill(start_color="5B7FA6", end_color="5B7FA6", fill_type="solid")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style data rows with alternating colors and proper number formatting
        for row in range(4, ws.max_row + 1):
            fill_color = "F2F2F2" if row % 2 == 0 else "FFFFFF"
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Special formatting for specific columns
                if col == 6:  # Profit column
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value >= 1000 or cell.value <= -1000:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '0.00'
                elif col in [1, 4]:  # Login and Deal ID columns - no thousands separator
                    if cell.value and isinstance(cell.value, int):
                        cell.number_format = '0'  # No thousands separator
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col == 2:  # Year column
                    if cell.value and isinstance(cell.value, int):
                        cell.number_format = '0'  # No thousands separator for year
                        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Add borders to all data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(3, ws.max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        # Set column widths - added agent and zip columns
        column_widths = [12, 8, 12, 12, 12, 15, 50, 20, 20, 12]  # Login, Year, Month, Deal ID, Category, Profit, Comment, Date, Agent, ZIP
        for i, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[column_letter].width = width
    
    def _get_login_group_mapping(self, database: str, logins: List[int]) -> Dict[int, str]:
        """Get login-group mapping from database"""
        try:
            from database_manager import DatabaseManager
            
            db = DatabaseManager()
            success = db.connect_to_database(database)
            if not success:
                return {}
            
            connection = db.connection
            cursor = connection.cursor()
            
            # Convert logins to a comma-separated string for IN clause
            login_list = ','.join(str(login) for login in logins)
            
            # Get login-group mapping
            query = f"SELECT Login, `Group` FROM mt5_users WHERE Login IN ({login_list})"
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # Build mapping dictionary
            mapping = {}
            for row in rows:
                mapping[int(row[0])] = row[1]
            
            cursor.close()
            db.close_connection()
            
            return mapping
            
        except Exception as e:
            print(f"❌ Error getting login-group mapping: {e}")
            return {}
    
    def _clean_group_data(self, value: str) -> str:
        """Clean group data to fix backslash and character issues"""
        if not isinstance(value, str):
            return value
        
        # Fix double backslashes to single backslashes
        cleaned = value.replace('\\\\', '\\')
        
        # Strip leading/trailing whitespace
        cleaned = cleaned.strip()
        
        # Remove any null characters or control characters
        cleaned = ''.join(char for char in cleaned if ord(char) >= 32 or char in '\t\n\r')
        
        return cleaned

    def _clean_withdrawal_value(self, value: str) -> str:
        """Clean withdrawal values and make them negative"""
        if not isinstance(value, str):
            return value
        
        # Remove currency symbols and clean the value
        cleaned = value.replace('$', '').replace(',', '').strip()
        
        # Try to convert to number and make negative
        try:
            num_value = float(cleaned)
            # Make withdrawals negative (if they're positive)
            if num_value > 0:
                return f"-{num_value:,.2f}"
            else:
                return f"{num_value:,.2f}"
        except (ValueError, TypeError):
            # If not a number, return as is
            return value

    def _clean_cell_data(self, cells: List[str]) -> List[str]:
        """Clean cell data including group column fixes and withdrawal formatting"""
        cleaned_cells = []
        
        # Check if this is a header row to identify withdrawal columns
        is_header_row = any(header in str(cell).lower() for cell in cells 
                           for header in ['login', 'group', 'deposits', 'withdrawals', 'net p/l'])
        
        for i, cell in enumerate(cells):
            if isinstance(cell, str):
                cleaned = cell.strip()
                
                # Group column cleaning (usually index 2)
                if i == 2:
                    cleaned = self._clean_group_data(cleaned)
                
                # Withdrawal column cleaning - check if header contains 'withdrawal'
                elif not is_header_row and i < len(cells):
                    # Look for withdrawal-related headers in this position
                    if any('withdrawal' in str(h).lower() for h in [cells[i]] if h):
                        cleaned = self._clean_withdrawal_value(cleaned)
                
                cleaned_cells.append(cleaned)
            else:
                cleaned_cells.append(cell)
        
        return cleaned_cells
    
    def _clean_cell_data_minimal(self, cells: List[str]) -> List[str]:
        """Minimal cell cleaning - only basic formatting with group column special handling"""
        cleaned_cells = []
        
        for i, cell in enumerate(cells):
            if cell is None:
                cleaned_cells.append('')
            elif isinstance(cell, str):
                # Only basic cleaning - remove excessive whitespace and null characters
                cleaned = cell.strip().replace('\x00', '').replace('\r', '')
                
                # Special handling for Group column (typically column 3, index 2)
                if i == 2:
                    cleaned = self._clean_group_data(cleaned)
                
                cleaned_cells.append(cleaned)
            else:
                cleaned_cells.append(str(cell) if cell is not None else '')
        
        return cleaned_cells
